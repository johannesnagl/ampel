#!/usr/bin/env python3
"""Convert "data/dishes 2.0.xlsx" → data/dishes.json.

Use after editing the catalog spreadsheet to regenerate the file the app loads.

Usage:
    python3 scripts/xlsx-to-json.py
"""
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

XLSX = Path("data/dishes 2.0.xlsx")
JSON_OUT = Path("data/dishes.json")
VERSION = 8

VALID_CATEGORIES = {"green", "yellow", "red"}
VALID_FREQ_TYPES = {"weekly", "monthly"}
VALID_SLOT_TYPES = {"breakfast", "lunch", "dinner", "snack", "dessert"}
VALID_TAGS = {"leicht verdaulich", "bowl", "süß", "warm", "kalt",
              "meal prep", "to go", "vegetarisch", "dessert", "cheat"}

def autotags(name: str) -> list[str]:
    tokens = re.findall(r"[A-Za-zÄÖÜäöüß]+", name.lower())
    return [t for t in tokens if len(t) > 2]

def parse_csv(value) -> list[str]:
    if value is None: return []
    return [s.strip() for s in str(value).split(",") if s.strip()]

def main():
    wb = load_workbook(XLSX)
    ws = wb["Dishes"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    dishes = []
    errors = []
    seen_ids = set()

    for r_idx, row in enumerate(rows, start=2):
        if row is None or row[0] is None or str(row[0]).strip() == "":
            continue
        id_, name, category, heavy, freq_type, freq_max, slot_types, tags, notes = row[:9]
        typ = row[9] if len(row) > 9 else None

        # Validate
        if id_ in seen_ids:
            errors.append(f"Row {r_idx}: duplicate id '{id_}'")
        seen_ids.add(id_)
        if category not in VALID_CATEGORIES:
            errors.append(f"Row {r_idx} ({id_}): invalid category '{category}'")
        if freq_type not in VALID_FREQ_TYPES:
            errors.append(f"Row {r_idx} ({id_}): invalid frequency_type '{freq_type}'")
        try:
            freq_max_n = int(freq_max)
        except (TypeError, ValueError):
            errors.append(f"Row {r_idx} ({id_}): invalid frequency_max '{freq_max}'")
            freq_max_n = 1
        slot_list = parse_csv(slot_types)
        invalid_slots = [s for s in slot_list if s not in VALID_SLOT_TYPES]
        if invalid_slots:
            errors.append(f"Row {r_idx} ({id_}): invalid slot_types {invalid_slots}")
        if not slot_list:
            errors.append(f"Row {r_idx} ({id_}): empty slot_types")

        heavy_bool = bool(heavy) if heavy not in (None, "", "FALSE", "false") else False
        if category == "green" and heavy_bool:
            heavy_bool = False  # enforce: green can't be heavy

        # Spalte H enthält die kuratierten semantischen Tags
        # (leicht verdaulich, bowl, süß, warm, kalt, meal prep, to go,
        #  vegetarisch, dessert, cheat). Nur wenn leer, aus dem Namen ableiten.
        tag_list = parse_csv(tags)
        if tag_list:
            unknown = [t for t in tag_list if t not in VALID_TAGS]
            if unknown:
                errors.append(f"Row {r_idx} ({id_}): unknown tag(s) {unknown}; "
                              f"allowed: {sorted(VALID_TAGS)}")
        elif name:
            tag_list = autotags(name)   # Fallback für Zeilen ohne kuratierte Tags

        dishes.append({
            "id": str(id_),
            "name": str(name),
            "category": category,
            "heavy": heavy_bool,
            "frequency": {"type": freq_type, "max": freq_max_n},
            "slotTypes": slot_list,
            "tags": tag_list,
            "notes": str(notes or ""),
            **({"typ": str(typ)} if typ else {}),
        })

    if errors:
        print("Validation errors:", file=sys.stderr)
        for e in errors:
            print(f"  {e}", file=sys.stderr)
        sys.exit(1)

    out = {"version": VERSION, "dishes": dishes}
    JSON_OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2))
    print(f"✓ {JSON_OUT} ({len(dishes)} dishes)")

if __name__ == "__main__":
    main()
